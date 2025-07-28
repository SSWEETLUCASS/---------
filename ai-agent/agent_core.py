import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from docx import Document
from docx.shared import Pt
from docx.enum.text import WD_ALIGN_PARAGRAPH
from gigachat_wrapper import get_llm
import re

from difflib import SequenceMatcher

def retrieve_similar_ideas(user_input: str, agents_data: list[str], threshold: float = 0.3) -> list[str]:
    """
    Простейшая реализация RAG: ищем инициативы, похожие по тексту на user_input
    """
    similar = []
    for idea in agents_data:
        ratio = SequenceMatcher(None, user_input.lower(), idea.lower()).ratio()
        if ratio > threshold:
            similar.append(idea)
    return similar

def check_idea_with_gigachat_local(user_input: str, user_data: dict, is_free_form: bool = False) -> tuple[str, bool, dict]:
    try:
        wb = load_workbook("agents.xlsx", data_only=True)
        ws = wb.active
        all_agents_data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[4]:
                continue

            block, ssp, owner, contact, name, short_name, desc, typ = row
            full_info = f"""Блок: {block}
ССП: {ssp}
Владелец: {owner}
Контакт: {contact}
Название инициативы: {name}
Краткое название: {short_name}
Описание: {desc}
Тип: {typ}"""
            all_agents_data.append(full_info)

        joined_data = "\n\n".join(all_agents_data) if all_agents_data else "(список инициатив пуст)"
    except Exception as e:
        print(f"⚠️ Ошибка при загрузке agents.xlsx: {e}")
        joined_data = "(не удалось загрузить данные об инициативах)"

    # 🎯 RAG — выбираем только похожие инициативы
    rag_context = retrieve_similar_ideas(user_input, all_agents_data)
    rag_context_text = "\n\n".join(rag_context) if rag_context else "Ничего похожего не найдено."

    if is_free_form:
        prompt = f"""
Вот список похожих инициатив (RAG):
{rag_context_text}

1. Проанализируй текст и заполни шаблон:
"Название", "Что хотим улучшить?", "Какие данные поступают агенту на выход?",
"Как процесс выглядит сейчас? as-is", "Какой результат нужен от агента?",
"Достижимый идеал(to-be)", "Масштаб процесса"

Если что-то не указано — скажи об этом.

Текст пользователя:
\"\"\"{user_data['Описание в свободной форме']}`\"\"\"

2. Сравни инициативу с найденными:
- Если идея похожа — "НЕ уникальна + название и владелец"
- Если новая — "Уникальна", предложи улучшения
- Если непонятно — "Извините, но я вас не понимаю"
"""
    else:
        prompt = f"""
Вот инициатива от пользователя:
Название: {user_data['Название инициативы']}
Что хотим улучшить?: {user_data['Что хотим улучшить?']}
Какие данные поступают агенту на выход?: {user_data['Какие данные поступают агенту на выход?']}
Как процесс выглядит сейчас? as-is: {user_data['Как процесс выглядит сейчас? as-is']}
Какой результат нужен от агента?: {user_data['Какой результат нужен от агента?']}
Достижимый идеал(to-be): {user_data['Достижимый идеал(to-be)']}
Масштаб процесса: {user_data['Масштаб процесса']}

Похожие инициативы (RAG):
{rag_context_text}

Сравни инициативу с ними, и прими решение: уникальна или нет?
"""

    raw_response = get_llm().invoke(prompt)
    response_text = str(raw_response).strip()

    is_unique = "уникальна" in response_text.lower() and "не уникальна" not in response_text.lower()

    parsed_data = {}
    if is_free_form:
        fields = [
            "Название", "Что хотим улучшить?", "Какие данные поступают агенту на выход?",
            "Как процесс выглядит сейчас? as-is", "Какой результат нужен от агента?",
            "Достижимый идеал(to-be)", "Масштаб процесса"
        ]
        for field in fields:
            match = re.search(rf"{field}[:\-–]\s*(.+)", response_text, re.IGNORECASE)
            if match:
                parsed_data[field] = match.group(1).strip()

    return response_text, is_unique, parsed_data


if __name__ == "__main__":
    while True:
        print("\nВведите данные инициативы (или 'выход'):")
        title = input("Название инициативы: ").strip()
        if title.lower() in ("выход", "exit", "quit"):
            break

        choice = input("❓Заполнить по шаблону или описать в свободной форме? (шаблон / свободно): ").strip().lower()

        if choice.startswith("шаб"):
            user_data = {
                "Название инициативы": title,
                "Что хотим улучшить?": input("Что хотим улучшить?: ").strip(),
                "Какие данные поступают агенту на выход?": input("Какие данные поступают агенту на выход?: ").strip(),
                "Как процесс выглядит сейчас? as-is": input("Как процесс выглядит сейчас? as-is: ").strip(),
                "Какой результат нужен от агента?": input("Какой результат нужен от агента?: ").strip(),
                "Достижимый идеал(to-be)": input("Достижимый идеал(to-be): ").strip(),
                "Масштаб процесса": input("Масштаб процесса: ").strip()
            }

            print("\n🔍 Проверка уникальности через GigaChat...")
            result, is_unique, parsed_data = check_idea_with_gigachat_local(title, user_data)

        else:
            free_text = input("📝 Опишите инициативу в свободной форме: ").strip()
            user_data = {"Описание в свободной форме": free_text}

            print("\n🔍 Проверка и структурирование через GigaChat...")
            result, is_unique, parsed_data = check_idea_with_gigachat_local(title, user_data, is_free_form=True)

        print("\n🧠 Ответ GigaChat:")
        match = re.search(r"content\s*=\s*['\"](.+?)['\"]", result)
        if match:
            print(match.group(1))
        else:
            print(result)

        if is_unique:
            print("\n✅ Идея уникальна!")

            if parsed_data:
                choice = input("📄 Сгенерировать шаблоны документов на основе распознанной информации? (да/нет): ").strip().lower()
                if choice in ("да", "д", "yes", "y"):
                    word_path, excel_path = generate_files(parsed_data)
                    print(f"\n📄 Файлы созданы:\n - {word_path}\n - {excel_path}")
                else:
                    print("🚫 Генерация отменена.")
            else:
                print("⚠️ Не удалось автоматически разобрать текст. Шаблоны не созданы.")

        else:
            print("\n⚠️ Идея не уникальна или неполна.")
