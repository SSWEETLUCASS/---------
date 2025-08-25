import os
import json
import logging
import re
from dotenv import load_dotenv
from dialog_bot_sdk.bot import DialogBot
from dialog_bot_sdk.entities.messaging import UpdateMessage, MessageContentType
from dialog_bot_sdk.entities.messaging import MessageHandler, CommandHandler
from openpyxl import load_workbook, Workbook

from ai_agent import (
    check_general_message_with_gigachat,
    check_idea_with_gigachat_local,
    generate_files,
    generate_agents_summary_file,
    find_agent_owners,
    generate_idea_suggestions,
    calculate_work_cost_interactive,
    generate_idea_evaluation_diagram,
    # Импорты для системы уточнений
    generate_cost_questions,
    process_cost_answers,
    calculate_final_cost,
)

# Загрузка конфигурации
with open('config.json', 'r', encoding='utf-8') as f:
    config = json.load(f)

# Загрузка переменных окружения
load_dotenv()

# Установка путей к сертификатам
os.environ["REQUESTS_CA_BUNDLE"] = config['file_settings']['certificates']['requests_ca_bundle']
os.environ["GRPC_DEFAULT_SSL_ROOTS_FILE_PATH"] = config['file_settings']['certificates']['grpc_roots']

BOT_TOKEN = os.getenv("DIALOG_BOT_TOKEN")

# Настройка логирования
logging.basicConfig(
    level=config['logging']['level'],
    format=config['logging']['format'],
    filename=config['logging']['file']
)

# Глобальные переменные
user_states = {}
bot = None

def send_file(peer, file_path, text=None, name=None):
    """Отправка файла с возможным описанием"""
    try:
        logging.info(f"📤 Отправка файла: {file_path}")
        with open(file_path, "rb") as f:
            bot.messaging.send_file_sync(
                peer,
                f,
                name=name or os.path.basename(file_path),
                caption=text or ""
            )
        return True
    except Exception as e:
        logging.error(f"❌ Ошибка отправки файла {file_path}: {e}")
        return False

def send_image(peer, image_path, caption=None):
    """Отправка изображения через бота"""
    try:
        logging.info(f"📤 Отправка изображения: {image_path}")
        with open(image_path, "rb") as f:
            bot.messaging.send_file_sync(
                peer,
                f,
                name=os.path.basename(image_path),
                caption=caption or ""
            )
        return True
    except Exception as e:
        logging.error(f"❌ Ошибка отправки изображения {image_path}: {e}")
        return False

def start_handler(update: UpdateMessage):
    user_id = update.peer.id
    user_states[user_id] = {"mode": config['states']['main_menu']}
    bot.messaging.send_message(update.peer, config['bot_settings']['commands']['start']['response'])

def idea_handler(update: UpdateMessage):
    peer = update.peer
    user_id = peer.id
    # Безопасное извлечение текста
    user_message = getattr(update.message, "text", "").strip()

    current_state = user_states.get(user_id, {})

    # Если уже в процессе работы над идеей
    if current_state.get("mode") in [
        config['states']['idea_template'],
        config['states']['idea_free_form'],
        "cost_questions",
        "awaiting_detailed_cost_decision"
    ]:
        bot.messaging.send_message(peer, "✍️ Продолжаем доработку вашей идеи...")
        return

    # Поля шаблона
    template_fields = [
        "Название",
        "Что хотим улучшить?",
        "Какие данные поступают агенту на выход?",
        "Как процесс выглядит сейчас? as-is",
        "Какой результат нужен от агента?",
        "Достижимый идеал(to-be)",
        "Масштаб процесса"
    ]

    # Проверка заполненности идеи
    def check_completeness(text: str) -> tuple[int, dict]:
        idea_data = {}
        matches = 0
        text_lower = text.lower()

        for field in template_fields:
            key_words = field.lower().replace("?", "").replace("(", "").replace(")", "").split()
            if any(kw in text_lower for kw in key_words):
                matches += 1
                idea_data[field] = f"(Найдено в тексте) {text}"
            else:
                idea_data[field] = ""

        return matches, idea_data

    matches, idea_data = check_completeness(user_message)

    if matches >= 5:
        # Почти полная идея → сразу формируем шаблон
        user_states[user_id] = {
            "mode": config['states']['idea_template'],
            "idea_data": idea_data
        }

        # Формируем красивый вывод
        template_text = "✅ Я собрал вашу идею в шаблон:\n\n"
        for field, value in idea_data.items():
            template_text += f"🔹 {field}: {value if value else '—'}\n"

        bot.messaging.send_message(peer, template_text)

        # 🚀 Сразу же передаём в AI-агент для оценки
        try:
            ai_prompt = ai_agent._generate_idea_prompt(
                joined_data="",  # сюда можно подставить базу существующих идей
                user_data=idea_data,
                is_free_form=False  # так как шаблон уже собран
            )
            ai_response = ai_agent.ask(ai_prompt)
            bot.messaging.send_message(peer, ai_response)
        except Exception as e:
            logging.error(f"Ошибка при передаче идеи в AI-агент: {e}")
            bot.messaging.send_message(peer, "⚠️ Не удалось обработать идею в AI-агенте.")
    else:
        # Идея не полная → переходим к пошаговому уточнению
        user_states[user_id] = {
            "mode": config['states']['idea_choose_format'],
            "current_field": 0,
            "idea_data": {"raw_text": user_message}
        }
        bot.messaging.send_message(peer, config['bot_settings']['commands']['idea']['responses']['initial'])




def agent_handler(update: UpdateMessage):
    peer = update.peer
    try:
        agents_file_path = config['file_settings']['agents_file']
        if not os.path.exists(agents_file_path):
            bot.messaging.send_message(peer, config['bot_settings']['commands']['ai_agent']['responses']['file_not_found'])
            wb = Workbook()
            ws = wb.active
            ws.append(["Блок", "ССП", "Владелец", "Контакт", "Название", "Краткое название", "Описание", "Тип"])
            wb.save(agents_file_path)

        summary_file = generate_agents_summary_file(agents_file_path)
        bot.messaging.send_message(peer, config['bot_settings']['commands']['ai_agent']['responses']['initial'])

        if not send_file(peer, agents_file_path):
            bot.messaging.send_message(peer, config['bot_settings']['commands']['ai_agent']['responses']['file_error'].format(file_type="основной"))

        if summary_file and os.path.exists(summary_file):
            if not send_file(peer, summary_file, text="📊 Аналитический отчет"):
                bot.messaging.send_message(peer, config['bot_settings']['commands']['ai_agent']['responses']['file_error'].format(file_type="аналитический"))
            try:
                os.remove(summary_file)
            except Exception as e:
                logging.warning(f"Не удалось удалить временный файл: {e}")

    except Exception as e:
        logging.error(f"Ошибка в agent_handler: {e}")
        bot.messaging.send_message(peer, config['error_messages']['file_error'].format(error=e))

def search_owners_handler(update: UpdateMessage):
    peer = update.peer
    user_id = peer.id
    query = update.message.text.strip() if update.message and update.message.text else ""

    try:
        agents_file_path = config['file_settings']['agents_file']
        if not os.path.exists(agents_file_path):
            bot.messaging.send_message(peer, config['error_messages']['file_not_found'])
            return

        wb = load_workbook(agents_file_path)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        agents_data = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]

        # Фильтрация по запросу
        results = []
        if query:
            query_lower = query.lower()
            for agent in agents_data:
                if any(query_lower in str(value).lower() for value in agent.values() if value):
                    results.append(agent)

        if results:
            reply = "🔎 Найдено совпадений: {}\n\n".format(len(results))
            for idx, agent in enumerate(results, start=1):
                reply += f"👤 {idx}. {agent.get('Name', 'Без имени')}\n"
                reply += f"📌 Описание: {agent.get('Description', '—')}\n"
                reply += f"🏷 Теги: {agent.get('Tags', '—')}\n\n"
        else:
            reply = "❌ Ничего не найдено по вашему запросу."

        # Сохраняем состояние
        user_states[user_id] = {
            "mode": config['states']['search_owners'],
            "agents_data": agents_data
        }

        bot.messaging.send_message(peer, reply)

    except Exception as e:
        logging.error(f"Ошибка в search_owners_handler: {e}")
        bot.messaging.send_message(peer, config['error_messages']['general_error'].format(error=e))


def consultation_handler(update: UpdateMessage):
    peer = update.peer
    # Consultation теперь = полезные ссылки
    links = config['bot_settings']['commands']['consultation']['responses']['links']
    bot.messaging.send_message(peer, f"📚 Полезные материалы:\n\n{links}")
    user_states[peer.id] = {"mode": config['states']['main_menu']}

def help_handler(update: UpdateMessage):
    bot.messaging.send_message(update.peer, config['bot_settings']['commands']['help']['response'])

def process_template_idea(update: UpdateMessage, user_id: int):
    peer = update.peer
    text = update.message.text_message.text.strip()
    state = user_states[user_id]
    current_field = state["current_field"]

    if current_field > 0:
        field_name = config['template_fields'][current_field - 1]
        state["idea_data"][field_name] = text

    if current_field < len(config['template_fields']):
        field_name = config['template_fields'][current_field]
        bot.messaging.send_message(peer, config['bot_settings']['commands']['idea']['responses']['template_field'].format(field=field_name))
        state["current_field"] += 1
    else:
        finalize_idea_analysis(peer, user_id, state, text, is_template=True)

def finalize_idea_analysis(peer, user_id, state, text, is_template=False):
    """Завершает анализ идеи и предлагает детальный расчет стоимости"""
    bot.messaging.send_message(peer, config['bot_settings']['commands']['idea']['responses']['complete'])
    
    try:
        state["idea_data"]["user_id"] = user_id
        response, is_unique, parsed_data, _ = check_idea_with_gigachat_local(
            text, state["idea_data"], is_free_form=not is_template
        )
        
        # Базовый расчет стоимости
        basic_cost_info = calculate_work_cost_interactive(parsed_data or state["idea_data"], is_unique)
        
        # Генерация и отправка диаграммы
        try:
            diagram_path = generate_idea_evaluation_diagram(state["idea_data"], is_unique, parsed_data)
            if diagram_path and os.path.exists(diagram_path):
                logging.info(f"📊 Отправка диаграммы оценки: {diagram_path}")
                send_image(peer, diagram_path, "📊 Диаграмма оценки идеи")
                try:
                    os.remove(diagram_path)
                    logging.info(f"🗑️ Временный файл диаграммы удален: {diagram_path}")
                except Exception as cleanup_error:
                    logging.warning(f"Не удалось удалить файл диаграммы: {cleanup_error}")
        except Exception as diagram_error:
            logging.error(f"Ошибка при создании диаграммы: {diagram_error}")
        
        # Отправляем результат анализа
        analysis_message = f"🧠 **Результат анализа:**\n\n{response}\n\n{basic_cost_info}"
        bot.messaging.send_message(peer, analysis_message)
        
        # Предлагаем детальный расчет
        detailed_cost_offer = (
            "💰 **Хотите получить детальный расчет стоимости?**\n\n"
            "📝 Я могу задать несколько уточняющих вопросов и сделать более точный расчет "
            "с разбивкой по этапам, команде и временным рамкам.\n\n"
            "✅ Напишите 'да' или 'детальный расчет' для продолжения\n"
            "❌ Или любое другое сообщение для завершения"
        )
        bot.messaging.send_message(peer, detailed_cost_offer)
        
        # Переводим в режим ожидания решения о детальном расчете
        user_states[user_id] = {
            "mode": "awaiting_detailed_cost_decision",
            "idea_data": parsed_data or state["idea_data"],
            "is_unique": is_unique,
            "basic_cost": basic_cost_info
        }
        
        # Генерируем файлы с базовой информацией
        if state["idea_data"]:
            try:
                word_path, excel_path = generate_files(state["idea_data"], basic_cost_info)
                bot.messaging.send_message(peer, config['bot_settings']['commands']['idea']['responses']['files_ready'])
                send_file(peer, word_path, text="📄 Техническое описание")
                send_file(peer, excel_path, text="📊 Структурированные данные")
                try:
                    os.remove(word_path)
                    os.remove(excel_path)
                except:
                    pass
            except Exception as file_error:
                logging.error(f"Ошибка при создании файлов: {file_error}")
                bot.messaging.send_message(peer, "⚠️ Файлы создать не удалось, но анализ завершен")

    except Exception as e:
        logging.error(f"Ошибка при обработке идеи: {e}")
        bot.messaging.send_message(peer, config['error_messages']['analysis_error'].format(error=e))
        user_states[user_id] = {"mode": config['states']['main_menu']}

def handle_cost_questions_mode(update: UpdateMessage, user_id: int):
    """Исправленная обработка режима уточняющих вопросов для расчета стоимости"""
    peer = update.peer
    text = update.message.text_message.text.strip()
    state = user_states[user_id]
    
    try:
        if state["mode"] == "cost_questions":
            # Пользователь отвечает на уточняющие вопросы
            questions = state.get("cost_questions", {})
            
            # Проверяем команды завершения
            finish_keywords = ['рассчитать', 'посчитать', 'готово', 'хватит', 'стоп', 'финиш', 'расчет']
            if any(word in text.lower() for word in finish_keywords):
                # Собираем уже данные ответы
                answers = {}
                for q_id, q_data in questions.items():
                    if q_data.get('answered', False) and q_data.get('answer'):
                        answers[q_id] = q_data['answer']
                
                if answers:
                    bot.messaging.send_message(peer, "⏳ Делаю финальный расчет стоимости на основе ваших ответов...")
                    final_cost, _ = calculate_final_cost(state["idea_data"], answers, user_id)
                    bot.messaging.send_message(peer, final_cost)
                    user_states[user_id] = {"mode": config['states']['main_menu']}
                    return
                else:
                    bot.messaging.send_message(peer, "❌ Нет ответов для расчета. Пожалуйста, ответьте хотя бы на несколько вопросов.")
                    return
            
            # Обрабатываем ответы пользователя
            updated_questions, all_answered, status_msg = process_cost_answers(questions, text)
            state["cost_questions"] = updated_questions
            
            bot.messaging.send_message(peer, status_msg)
            
            if all_answered:
                # Все ответы получены, делаем финальный расчет
                bot.messaging.send_message(peer, "⏳ Все ответы получены! Делаю детальный расчет...")
                answers = {}
                for q_id, q_data in updated_questions.items():
                    if q_data.get('answer'):
                        answers[q_id] = q_data['answer']
                
                final_cost, _ = calculate_final_cost(state["idea_data"], answers, user_id)
                bot.messaging.send_message(peer, final_cost)
                user_states[user_id] = {"mode": config['states']['main_menu']}
            
        elif state["mode"] == "awaiting_detailed_cost_decision":
            # Пользователь решает, нужен ли детальный расчет
            positive_keywords = ['да', 'детальный', 'расчет', 'уточнения', 'вопросы', 'точный', 'подробный']
            if any(word in text.lower() for word in positive_keywords):
                bot.messaging.send_message(peer, "⏳ Генерирую уточняющие вопросы для точного расчета...")
                
                # Генерируем вопросы для уточнения
                questions_text, questions_dict = generate_cost_questions(state["idea_data"])
                
                if questions_dict and questions_text:
                    bot.messaging.send_message(peer, questions_text)
                    user_states[user_id] = {
                        "mode": "cost_questions",
                        "idea_data": state["idea_data"],
                        "cost_questions": questions_dict,
                        "is_unique": state.get("is_unique", True)
                    }
                    logging.info(f"[User {user_id}] Переведен в режим cost_questions с {len(questions_dict)} вопросами")
                else:
                    bot.messaging.send_message(peer, "⚠️ Не удалось сгенерировать вопросы. Используем базовый расчет.")
                    user_states[user_id] = {"mode": config['states']['main_menu']}
            else:
                # Пользователь не хочет детальный расчет
                bot.messaging.send_message(peer, "✅ Понятно! Базовый расчет стоимости уже предоставлен выше.")
                user_states[user_id] = {"mode": config['states']['main_menu']}
                
    except Exception as e:
        logging.error(f"Ошибка в обработке вопросов стоимости: {e}")
        bot.messaging.send_message(peer, f"⚠️ Произошла ошибка: {e}")
        user_states[user_id] = {"mode": config['states']['main_menu']}

def text_handler(update: UpdateMessage, widget=None):
    if not update.message or not update.message.text_message:
        return
    text = update.message.text_message.text.strip()
    user_id = update.peer.id
    peer = update.peer
    state = user_states.get(user_id, {"mode": config['states']['main_menu']})

    # Логирование для отладки
    logging.info(f"[User {user_id}] Message: {text[:100]}... | Mode: {state.get('mode', 'none')}")

    # === ОБРАБОТКА РЕЖИМОВ РАСЧЕТА СТОИМОСТИ (ПРИОРИТЕТ) ===
    if state.get("mode") in ["cost_questions", "awaiting_detailed_cost_decision"]:
        handle_cost_questions_mode(update, user_id)
        return

    # === ОБРАБОТКА ДРУГИХ СПЕЦИАЛЬНЫХ РЕЖИМОВ ===
    if state.get("mode") == config['states']['idea_choose_format']:
        if "шаблон" in text.lower():
            state["mode"] = config['states']['idea_template']
            state["current_field"] = 0
            state["idea_data"] = {}
            process_template_idea(update, user_id)
        elif "сам" in text.lower() or "свобод" in text.lower():
            state["mode"] = config['states']['idea_free_form']
            bot.messaging.send_message(peer, config['bot_settings']['commands']['idea']['responses']['free_form_prompt'])
        else:
            bot.messaging.send_message(peer, config['bot_settings']['commands']['idea']['responses']['template_choice_error'])
        return

    elif state.get("mode") == config['states']['idea_template']:
        process_template_idea(update, user_id)
        return

    # === Если уже в режиме работы с идеей ===
    if state.get("mode") in [config['states']['idea_template'], config['states']['idea_free_form']]:
        if state["mode"] == config['states']['idea_template']:
            process_template_idea(update, user_id)
        elif state["mode"] == config['states']['idea_free_form']:
            user_data = {"Описание (уточнение)": text, "user_id": user_id}
            finalize_idea_analysis(peer, user_id, {"idea_data": user_data}, text, is_template=False)
        return

    elif state.get("mode") == config['states']['search_owners']:
        bot.messaging.send_message(peer, "🔍 Ищу подходящих владельцев...")
        try:
            owners_info = find_agent_owners(text)
            bot.messaging.send_message(peer, owners_info)
        except Exception as e:
            logging.error(f"Ошибка при поиске владельцев: {e}")
            bot.messaging.send_message(peer, config['error_messages']['general_error'].format(error=e))
        user_states[user_id] = {"mode": config['states']['main_menu']}
        return

    elif state.get("mode") == config['states']['help_with_ideas']:
        bot.messaging.send_message(peer, "💡 Генерирую идеи специально для вас...")
        try:
            ideas_response = generate_idea_suggestions(text)
            bot.messaging.send_message(peer, f"🎯 **Вот идеи для вас:**\n\n{ideas_response}")
        except Exception as e:
            logging.error(f"Ошибка при генерации идей: {e}")
            bot.messaging.send_message(peer, config['error_messages']['general_error'].format(error=e))
        user_states[user_id] = {"mode": config['states']['main_menu']}
        return

    # === ОБЫЧНЫЙ ДИАЛОГ ЧЕРЕЗ GIGACHAT ===
    try:
        logging.info(f"[User {user_id}] Sending to GigaChat with memory...")
        gpt_response, detected_command = check_general_message_with_gigachat(text, user_id)

        # Если в тексте GPT есть команда, но detected_command пуст
        if not detected_command and gpt_response:
            cmd_match = re.search(r"CMD:(\w+)", gpt_response, re.IGNORECASE)
            if cmd_match:
                detected_command = cmd_match.group(1).lower().strip()
                logging.info(f"[User {user_id}] Extracted command from GPT text: {detected_command}")

        if detected_command:
            logging.info(f"[User {user_id}] Detected command: {detected_command}")
            # Команды для выполнения
            command_map = {
                "start": start_handler,
                "ai_agent": agent_handler,
                "search_owners": search_owners_handler,
                "idea": idea_handler,
                "consultation": consultation_handler,
                "help": help_handler
            }
            handler = command_map.get(detected_command)
            if handler:
                # ИСПРАВЛЕНИЕ: Отправляем ответ GPT только если он не содержит команду или содержит полезную информацию
                if gpt_response and gpt_response.strip():
                    clean_gpt_response = re.sub(r'\s*CMD:\w+\s*', '', gpt_response).strip()
                    # Отправляем только если после очистки остался содержательный текст
                    if clean_gpt_response and len(clean_gpt_response) > 10:
                        bot.messaging.send_message(peer, clean_gpt_response)
                
                # Выполняем команду
                handler(update)
            else:
                logging.warning(f"[User {user_id}] No handler found for command: {detected_command}")
                bot.messaging.send_message(peer, gpt_response)
        else:
            # Обычный ответ без команды
            if gpt_response and gpt_response.strip():
                bot.messaging.send_message(peer, gpt_response)
                logging.info(f"[User {user_id}] Response sent successfully")
            else:
                fallback_msg = "🤔 Не совсем понял ваш вопрос. Попробуйте иначе или используйте /help"
                bot.messaging.send_message(peer, fallback_msg)
                logging.info(f"[User {user_id}] Fallback response sent")

    except Exception as e:
        error_msg = f"⚠️ Произошла ошибка при обработке сообщения: {str(e)}"
        logging.error(f"[User {user_id}] Error in text_handler: {e}")
        bot.messaging.send_message(peer, error_msg)

def main():
    global bot
    bot = DialogBot.create_bot({
        "endpoint": config['bot_settings']['endpoint'],
        "token": BOT_TOKEN,
        "is_secure": config['bot_settings']['is_secure'],
    })
    
    handlers = []
    
    # Основные команды из конфига
    for cmd, cmd_data in config['bot_settings']['commands'].items():
        handler_func = globals()[cmd_data['handler']]
        handlers.append(CommandHandler(handler_func, cmd))
        if 'aliases' in cmd_data:
            for alias in cmd_data['aliases']:
                handlers.append(CommandHandler(handler_func, alias))
    
    bot.messaging.command_handler(handlers)
    bot.messaging.message_handler([
        MessageHandler(text_handler, MessageContentType.TEXT_MESSAGE)
    ])
    
    logging.info("🤖 Бот запущен с поддержкой памяти диалогов!")
    logging.info("🧠 GigaChat будет автоматически помнить последние 10 сообщений каждого пользователя")
    logging.info("📊 Включена поддержка диаграмм оценки идей!")
    logging.info("💰 Включена исправленная система детального расчета стоимости!")
    logging.info("✅ Исправлена проблема дублирования ответов!")
    
    bot.updates.on_updates(do_read_message=True, do_register_commands=True)

if __name__ == "__main__":
    main()